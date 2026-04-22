from __future__ import annotations

from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl.styles import Font, PatternFill


app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
CATEGORIES = np.array(["房租", "餐饮", "交通", "通讯", "生活杂项"])
BASE_WEIGHTS = np.array([0.30, 0.20, 0.10, 0.05, 0.25], dtype=float)
MONTHS = np.array([f"{index}月" for index in range(1, 13)])
DEFAULT_SAVINGS_RATIO = 0.10


def validate_inputs(salary: float, savings_ratio: float) -> None:
    if salary <= 0:
        raise ValueError("月收入必须大于 0。")
    if salary > 1_000_000:
        raise ValueError("月收入过大，请输入合理范围内的金额。")
    if not 0 <= savings_ratio < 1:
        raise ValueError("储蓄比例必须在 0 到 100% 之间。")


def build_budget_plan(monthly_salary: float, savings_ratio: float) -> dict:
    validate_inputs(monthly_salary, savings_ratio)

    spending_ratio = 1 - savings_ratio
    distribution = BASE_WEIGHTS / np.sum(BASE_WEIGHTS)
    budget_amounts = monthly_salary * spending_ratio * distribution
    savings_amount = monthly_salary * savings_ratio
    remaining_amount = monthly_salary - np.sum(budget_amounts) - savings_amount
    annual_budget_matrix = np.tile(budget_amounts, (12, 1))

    items = [
        {
            "category": category,
            "ratio": round(float(spending_ratio * weight), 4),
            "percent": round(float(spending_ratio * weight * 100), 2),
            "amount": round(float(amount), 2),
        }
        for category, weight, amount in zip(CATEGORIES, distribution, budget_amounts)
    ]

    return {
        "monthly_salary": round(float(monthly_salary), 2),
        "savings_ratio": round(float(savings_ratio), 4),
        "savings_percent": round(float(savings_ratio * 100), 2),
        "savings_amount": round(float(savings_amount), 2),
        "remaining_amount": round(float(remaining_amount), 2),
        "items": items,
        "annual_budget_matrix": np.round(annual_budget_matrix, 2).tolist(),
    }


def build_budget_dataframe(plan: dict) -> pd.DataFrame:
    matrix = np.asarray(plan["annual_budget_matrix"], dtype=float)
    monthly_totals = np.sum(matrix, axis=1)
    savings_column = np.full(MONTHS.shape[0], plan["savings_amount"], dtype=float)
    salary_column = np.full(MONTHS.shape[0], plan["monthly_salary"], dtype=float)

    budget_df = pd.DataFrame(matrix, columns=CATEGORIES)
    budget_df.insert(0, "月份", MONTHS)
    budget_df["强制储蓄"] = np.round(savings_column, 2)
    budget_df["月度支出预算合计"] = np.round(monthly_totals, 2)
    budget_df["月收入"] = np.round(salary_column, 2)
    budget_df["预算结余校验"] = np.round(salary_column - monthly_totals - savings_column, 2)
    return budget_df


def build_actual_template(plan: dict) -> pd.DataFrame:
    records = []
    for month in MONTHS:
        for item in plan["items"]:
            records.append(
                {
                    "月份": month,
                    "类别": item["category"],
                    "预算金额": item["amount"],
                    "实际支出": 0.0,
                }
            )

    actual_df = pd.DataFrame(records)
    actual_df["差异"] = actual_df["预算金额"] - actual_df["实际支出"]
    actual_df["状态"] = np.where(actual_df["差异"] < 0, "超支", "结余")
    return actual_df


def write_excel_report(plan: dict, output: BytesIO) -> None:
    budget_df = build_budget_dataframe(plan)
    actual_df = build_actual_template(plan)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        budget_df.to_excel(writer, index=False, sheet_name="年度预算总表")
        actual_df.to_excel(writer, index=False, sheet_name="实际支出记录")

        budget_sheet = writer.sheets["年度预算总表"]
        actual_sheet = writer.sheets["实际支出记录"]

        header_fill = PatternFill(fill_type="solid", fgColor="DCEBDD")
        warning_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
        header_font = Font(bold=True)

        for sheet in (budget_sheet, actual_sheet):
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        for column in budget_sheet.columns:
            budget_sheet.column_dimensions[column[0].column_letter].width = 16

        for column in actual_sheet.columns:
            actual_sheet.column_dimensions[column[0].column_letter].width = 16

        budget_sheet.freeze_panes = "A2"
        actual_sheet.freeze_panes = "A2"

        for row_index in range(2, actual_sheet.max_row + 1):
            budget_cell = f"C{row_index}"
            actual_cell = f"D{row_index}"
            diff_cell = f"E{row_index}"
            status_cell = f"F{row_index}"
            actual_sheet[diff_cell] = f"={budget_cell}-{actual_cell}"
            actual_sheet[status_cell] = f'=IF({diff_cell}<0,"🔴超支","🟢结余")'
            if row_index % len(CATEGORIES) == 1:
                for col in range(1, 7):
                    actual_sheet.cell(row=row_index, column=col).fill = warning_fill

        output.seek(0)


@app.get("/")
def index():
    return render_template(
        "index.html",
        default_salary=6000,
        default_savings_ratio=round(DEFAULT_SAVINGS_RATIO * 100, 2),
    )


@app.post("/api/budget")
def calculate_budget():
    payload = request.get_json(silent=True) or {}

    try:
        salary = float(payload.get("salary", 0))
        savings_ratio = float(payload.get("savings_ratio", DEFAULT_SAVINGS_RATIO * 100)) / 100
        plan = build_budget_plan(salary, savings_ratio)
    except (TypeError, ValueError) as exc:
        return jsonify({"message": str(exc)}), 400

    return jsonify(
        {
            "message": "预算方案已生成。",
            "plan": plan,
        }
    )


@app.post("/api/export")
def export_budget():
    payload = request.get_json(silent=True) or {}

    try:
        salary = float(payload.get("salary", 0))
        savings_ratio = float(payload.get("savings_ratio", DEFAULT_SAVINGS_RATIO * 100)) / 100
        plan = build_budget_plan(salary, savings_ratio)
    except (TypeError, ValueError) as exc:
        return jsonify({"message": str(exc)}), 400

    file_buffer = BytesIO()
    write_excel_report(plan, file_buffer)

    file_name = (
        f"graduation_budget_{int(round(plan['monthly_salary']))}_"
        f"{int(round(plan['savings_percent']))}.xlsx"
    )
    output_path = BASE_DIR / file_name
    output_path.write_bytes(file_buffer.getvalue())
    file_buffer.seek(0)

    return send_file(
        file_buffer,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
