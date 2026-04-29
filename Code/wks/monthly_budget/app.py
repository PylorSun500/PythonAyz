from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from flask import Flask, render_template, request, send_file, session
from openpyxl.styles import Alignment, Font, PatternFill


app = Flask(__name__)
app.config["SECRET_KEY"] = "monthly-budget-demo-key"

BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

DEFAULT_BUDGET = 1100
MIN_BUDGET = 1000
MAX_BUDGET = 1200
TOTAL_LUNCHES = 30
WEEKS = 6
MEALS_PER_WEEK = 5
MIN_FISH_PER_WEEK = 2
OLIVE_OIL_NAME = "特级初榨橄榄油"
OLIVE_OIL_COST = 3.0
WEEKDAY_LABELS = np.array(["周一", "周二", "周三", "周四", "周五"], dtype=object)

PROTEIN_NAMES = np.array(
    ["三文鱼", "鳕鱼", "沙丁鱼", "金枪鱼", "鹰嘴豆", "鸡胸肉", "扁豆", "希腊酸奶烤鸡"],
    dtype=object,
)
PROTEIN_COSTS = np.array([16.0, 14.0, 11.0, 12.0, 8.0, 10.0, 8.5, 11.5], dtype=float)
PROTEIN_TYPES = np.array(
    ["鱼类", "鱼类", "鱼类", "鱼类", "豆类", "禽肉", "豆类", "禽肉"],
    dtype=object,
)

GRAIN_NAMES = np.array(["藜麦", "糙米", "全麦意面", "保加利亚小麦", "全麦库斯库斯"], dtype=object)
GRAIN_COSTS = np.array([8.0, 5.0, 6.0, 6.5, 6.0], dtype=float)

VEG_NAMES = np.array(["番茄", "黄瓜", "菠菜", "西兰花", "彩椒", "西葫芦", "茄子", "羽衣甘蓝"], dtype=object)
VEG_COSTS = np.array([3.0, 3.0, 4.0, 4.0, 4.0, 4.0, 5.0, 5.0], dtype=float)

EXTRA_NAMES = np.array(["菲达奶酪", "鹰嘴豆泥", "柠檬香草酱", "黑橄榄", "希腊酸奶酱"], dtype=object)
EXTRA_COSTS = np.array([7.0, 6.0, 6.0, 6.0, 6.5], dtype=float)


def validate_budget(monthly_budget: float) -> float:
    if monthly_budget < MIN_BUDGET or monthly_budget > MAX_BUDGET:
        raise ValueError(f"月度预算必须在 {MIN_BUDGET}-{MAX_BUDGET} 元之间。")
    return round(float(monthly_budget), 2)


def build_candidate_meals() -> pd.DataFrame:
    protein_idx, grain_idx, veg_a_idx, veg_b_idx, extra_idx = np.meshgrid(
        np.arange(PROTEIN_NAMES.size),
        np.arange(GRAIN_NAMES.size),
        np.arange(VEG_NAMES.size),
        np.arange(VEG_NAMES.size),
        np.arange(EXTRA_NAMES.size),
        indexing="ij",
    )

    protein_idx = protein_idx.ravel()
    grain_idx = grain_idx.ravel()
    veg_a_idx = veg_a_idx.ravel()
    veg_b_idx = veg_b_idx.ravel()
    extra_idx = extra_idx.ravel()

    different_vegetables = veg_a_idx != veg_b_idx
    total_costs = (
        PROTEIN_COSTS[protein_idx]
        + GRAIN_COSTS[grain_idx]
        + VEG_COSTS[veg_a_idx]
        + VEG_COSTS[veg_b_idx]
        + EXTRA_COSTS[extra_idx]
        + OLIVE_OIL_COST
    )

    valid_cost = (total_costs >= 30.0) & (total_costs <= 50.0)
    mask = different_vegetables & valid_cost

    protein_idx = protein_idx[mask]
    grain_idx = grain_idx[mask]
    veg_a_idx = veg_a_idx[mask]
    veg_b_idx = veg_b_idx[mask]
    extra_idx = extra_idx[mask]
    total_costs = total_costs[mask]

    lunch_names = np.array(
        [
            f"{PROTEIN_NAMES[p]} + {GRAIN_NAMES[g]} + {VEG_NAMES[v1]}/{VEG_NAMES[v2]} + {EXTRA_NAMES[e]}"
            for p, g, v1, v2, e in zip(protein_idx, grain_idx, veg_a_idx, veg_b_idx, extra_idx)
        ],
        dtype=object,
    )

    return pd.DataFrame(
        {
            "菜单名称": lunch_names,
            "主蛋白": PROTEIN_NAMES[protein_idx],
            "蛋白类别": PROTEIN_TYPES[protein_idx],
            "主食": GRAIN_NAMES[grain_idx],
            "蔬菜1": VEG_NAMES[veg_a_idx],
            "蔬菜2": VEG_NAMES[veg_b_idx],
            "配料": EXTRA_NAMES[extra_idx],
            "调味": OLIVE_OIL_NAME,
            "估算成本": np.round(total_costs, 2),
        }
    )


def choose_meals_from_pool(
    candidate_indices: np.ndarray,
    used_mask: np.ndarray,
    count: int,
    target_avg_cost: float,
    costs: np.ndarray,
    rng: np.random.Generator,
    tolerance: float,
) -> np.ndarray:
    available = candidate_indices[~used_mask[candidate_indices]]
    if available.size < count:
        raise ValueError("可选菜单数量不足，无法满足当前约束。")

    preferred = available[costs[available] <= max(30.0, target_avg_cost + tolerance)]
    pool = preferred if preferred.size >= count else available

    weights = 1.0 / (1.0 + np.abs(costs[pool] - target_avg_cost))
    weights = weights / np.sum(weights)
    return rng.choice(pool, size=count, replace=False, p=weights)


def generate_monthly_plan(monthly_budget: float) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    budget = validate_budget(monthly_budget)
    candidates = build_candidate_meals()
    all_indices = np.arange(len(candidates))
    fish_indices = np.flatnonzero(candidates["蛋白类别"].to_numpy() == "鱼类")
    costs = candidates["估算成本"].to_numpy(dtype=float)

    best_selection: list[int] | None = None
    best_gap = float("inf")

    for attempt in range(800):
        rng = np.random.default_rng(seed=int(budget * 100) + attempt * 37)
        used_mask = np.zeros(len(candidates), dtype=bool)
        selected_indices: list[int] = []
        remaining_budget = budget
        remaining_slots = TOTAL_LUNCHES
        success = True

        for _week in range(WEEKS):
            target_avg_cost = remaining_budget / max(remaining_slots, 1)

            try:
                weekly_fish = choose_meals_from_pool(
                    fish_indices,
                    used_mask,
                    MIN_FISH_PER_WEEK,
                    target_avg_cost,
                    costs,
                    rng,
                    tolerance=3.5,
                )
            except ValueError:
                success = False
                break

            used_mask[weekly_fish] = True
            selected_indices.extend(weekly_fish.tolist())
            remaining_budget -= float(np.sum(costs[weekly_fish]))
            remaining_slots -= MIN_FISH_PER_WEEK

            target_avg_cost = remaining_budget / max(remaining_slots, 1)
            try:
                weekly_other = choose_meals_from_pool(
                    all_indices,
                    used_mask,
                    MEALS_PER_WEEK - MIN_FISH_PER_WEEK,
                    target_avg_cost,
                    costs,
                    rng,
                    tolerance=3.0,
                )
            except ValueError:
                success = False
                break

            used_mask[weekly_other] = True
            selected_indices.extend(weekly_other.tolist())
            remaining_budget -= float(np.sum(costs[weekly_other]))
            remaining_slots -= MEALS_PER_WEEK - MIN_FISH_PER_WEEK

        if not success:
            continue

        total_cost = float(np.sum(costs[selected_indices]))
        if total_cost <= budget:
            gap = budget - total_cost
            if gap < best_gap:
                best_gap = gap
                best_selection = selected_indices
                if gap <= 5.0:
                    break

    if best_selection is None:
        raise ValueError("当前预算下未能生成满足约束的 30 顿午餐，请尝试提高预算。")

    meal_plan = candidates.iloc[best_selection].reset_index(drop=True).copy()
    meal_plan.insert(0, "周次", np.repeat(np.arange(1, WEEKS + 1), MEALS_PER_WEEK))
    meal_plan.insert(1, "星期", np.tile(WEEKDAY_LABELS, WEEKS))
    meal_plan.insert(2, "是否鱼类", np.where(meal_plan["蛋白类别"] == "鱼类", "是", "否"))

    shopping_list = build_shopping_list(meal_plan)
    summary = build_summary(meal_plan, shopping_list, budget)
    return meal_plan, shopping_list, summary


def build_shopping_list(meal_plan: pd.DataFrame) -> pd.DataFrame:
    ingredient_catalog = {
        **{
            name: {"类别": "主蛋白", "单份参考价": cost}
            for name, cost in zip(PROTEIN_NAMES.tolist(), PROTEIN_COSTS.tolist())
        },
        **{
            name: {"类别": "全谷物", "单份参考价": cost}
            for name, cost in zip(GRAIN_NAMES.tolist(), GRAIN_COSTS.tolist())
        },
        **{
            name: {"类别": "蔬菜", "单份参考价": cost}
            for name, cost in zip(VEG_NAMES.tolist(), VEG_COSTS.tolist())
        },
        **{
            name: {"类别": "配料", "单份参考价": cost}
            for name, cost in zip(EXTRA_NAMES.tolist(), EXTRA_COSTS.tolist())
        },
        OLIVE_OIL_NAME: {"类别": "调味", "单份参考价": OLIVE_OIL_COST},
    }

    ingredient_records: list[dict] = []
    for row in meal_plan.itertuples(index=False):
        ingredient_records.extend(
            [
                {"食材": row.主蛋白},
                {"食材": row.主食},
                {"食材": row.蔬菜1},
                {"食材": row.蔬菜2},
                {"食材": row.配料},
                {"食材": row.调味},
            ]
        )

    shopping_df = pd.DataFrame(ingredient_records)
    shopping_df["份数"] = 1
    shopping_df = shopping_df.groupby("食材", as_index=False)["份数"].sum()
    shopping_df["类别"] = shopping_df["食材"].map(lambda item: ingredient_catalog[item]["类别"])
    shopping_df["单份参考价"] = shopping_df["食材"].map(lambda item: ingredient_catalog[item]["单份参考价"])
    shopping_df["采购估算小计"] = np.round(shopping_df["份数"] * shopping_df["单份参考价"], 2)
    shopping_df = shopping_df[["类别", "食材", "份数", "单份参考价", "采购估算小计"]]
    records: list[dict[str, object]] = []
    for _, row in shopping_df.iterrows():
        records.append(
            {
                "类别": str(row["类别"]),
                "食材": str(row["食材"]),
                "份数": int(row["份数"]),
                "单份参考价": float(row["单份参考价"]),
                "采购估算小计": float(row["采购估算小计"]),
            }
        )

    return pd.DataFrame(records, columns=["类别", "食材", "份数", "单份参考价", "采购估算小计"])


def build_summary(meal_plan: pd.DataFrame, shopping_list: pd.DataFrame, monthly_budget: float) -> dict:
    weekly_fish = (
        meal_plan.groupby("周次")["蛋白类别"]
        .apply(lambda items: int(np.sum(np.asarray(items) == "鱼类")))
        .reset_index(name="鱼类次数")
    )
    total_cost = round(float(meal_plan["估算成本"].sum()), 2)
    return {
        "月度预算": round(float(monthly_budget), 2),
        "午餐总成本": total_cost,
        "预算结余": round(float(monthly_budget - total_cost), 2),
        "平均每餐成本": round(float(meal_plan["估算成本"].mean()), 2),
        "鱼类午餐次数": int(np.sum(meal_plan["蛋白类别"].to_numpy() == "鱼类")),
        "午餐数量": int(len(meal_plan)),
        "采购总计": round(float(shopping_list["采购估算小计"].sum()), 2),
        "每周鱼类统计": weekly_fish.to_dict(orient="records"),
    }


def serialize_result(meal_plan: pd.DataFrame, shopping_list: pd.DataFrame, summary: dict) -> dict:
    return {
        "meals": meal_plan.to_dict(orient="records"),
        "shopping": shopping_list.to_dict(orient="records"),
        "summary": summary,
    }


def export_to_excel(result: dict) -> tuple[BytesIO, str]:
    meal_plan = pd.DataFrame(result["meals"])
    shopping_list = pd.DataFrame(result["shopping"])
    summary = pd.DataFrame(
        [
            {"指标": key, "数值": value}
            for key, value in result["summary"].items()
            if key != "每周鱼类统计"
        ]
    )
    weekly_fish = pd.DataFrame(result["summary"]["每周鱼类统计"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        meal_plan.to_excel(writer, index=False, sheet_name="30天午餐计划")
        shopping_list.to_excel(writer, index=False, sheet_name="采购清单")
        summary.to_excel(writer, index=False, sheet_name="预算摘要")
        weekly_fish.to_excel(writer, index=False, sheet_name="预算摘要", startrow=len(summary) + 3)

        workbook = writer.book
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        header_font = Font(bold=True)

        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                adjusted_width = min(max(max_length + 2, 12), 30)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        summary_sheet = workbook["预算摘要"]
        summary_title_row = len(summary) + 4
        summary_sheet.cell(row=summary_title_row, column=1, value="每周鱼类统计")
        summary_sheet.cell(row=summary_title_row, column=1).font = Font(bold=True)

    output.seek(0)
    file_name = f"monthly_budget_plan_{int(result['summary']['月度预算'])}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, file_name


@app.get("/")
def index():
    result = session.get("latest_result")
    return render_template("index.html", default_budget=DEFAULT_BUDGET, result=result, error_message=None)


@app.post("/generate")
def generate():
    budget_text = request.form.get("monthly_budget", str(DEFAULT_BUDGET)).strip()

    try:
        monthly_budget = validate_budget(float(budget_text))
        meal_plan, shopping_list, summary = generate_monthly_plan(monthly_budget)
        result = serialize_result(meal_plan, shopping_list, summary)
        session["latest_result"] = result
        return render_template("index.html", default_budget=monthly_budget, result=result, error_message=None)
    except ValueError as exc:
        return render_template(
            "index.html",
            default_budget=budget_text or DEFAULT_BUDGET,
            result=session.get("latest_result"),
            error_message=str(exc),
        )


@app.post("/export")
def export():
    result = session.get("latest_result")
    if not result:
        return render_template(
            "index.html",
            default_budget=DEFAULT_BUDGET,
            result=None,
            error_message="请先生成一份午餐计划，再导出 Excel。",
        )

    output, file_name = export_to_excel(result)
    export_path = EXPORT_DIR / file_name
    export_path.write_bytes(output.getvalue())
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
